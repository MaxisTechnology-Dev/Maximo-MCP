"""
tools/reporting.py — KPI dashboards, Excel/PDF export, and Carbon table rendering.
"""

import base64
import io
import time
from datetime import datetime, timedelta
from typing import Any, Dict, List, Optional

from core.generic_oslc import query_object_structure
from core.maximo_client import MaximoAPIError, MaximoAuthError, get_connected_client
from core.oslc_utils import oslc_escape
from core.rbac import require_role


def _envelope(data: Any, cached: bool = False, duration_ms: int = 0) -> Dict:
    return {"success": True, "data": data, "metadata": {"cached": cached, "duration_ms": duration_ms}}


def _error(message: str, code: str = "API_ERROR") -> Dict:
    return {"success": False, "error": message, "error_code": code}


@require_role("manager")
async def get_maintenance_kpi_dashboard(
    site_id: str,
    period_months: int = 3,
) -> Dict[str, Any]:
    """
    Return a comprehensive maintenance KPI dashboard for a site.
    Includes MTTR, MTBF, PM compliance, cost trends, backlog, and technician utilisation.

    Args:
        site_id:       Site ID to analyse
        period_months: Analysis window in months (default: 3)

    Returns:
        Full KPI object with sub-sections for WO performance, PM compliance, cost, and labour.
    """
    if not site_id:
        return _error("site_id is required", "VALIDATION_ERROR")

    start = time.monotonic()
    cutoff = (datetime.now() - timedelta(days=period_months * 30)).strftime("%Y-%m-%dT00:00:00+00:00")
    now_str = datetime.now().strftime("%Y-%m-%dT%H:%M:%S+00:00")

    try:
        # --- Work orders ---
        # Single-condition WHERE only (siteid) — compound WHERE causes transport errors.
        # order_by="-reportdate" + page_size=50 fetches 50 most recent WOs; Python filters by date.
        wo_result = await query_object_structure(
            entity="workorder",
            filters={"site_id": site_id},
            select=["wonum", "status", "siteid", "reportdate", "actlabhrs", "targcompdate", "wopriority", "worktype", "actfinish", "schedfinish", "actlabcost"],
            order_by="-reportdate",
            page_size=50,
        )
        wos_raw: List[Dict] = wo_result.get("data", []) if "error" not in wo_result else []
        wos: List[Dict] = [w for w in wos_raw if w.get("reportdate", "") >= cutoff]

        completed = [w for w in wos if w.get("status") in ("COMP", "CLOSE")]
        corrective = [w for w in wos if w.get("worktype") in ("CM", "EM")]
        backlog = [w for w in wos if w.get("status") not in ("COMP", "CAN", "CLOSE")]
        overdue = [w for w in backlog if w.get("schedfinish", "9999") < now_str]

        # MTTR (mean time to repair) — hours
        repair_times = []
        for w in [x for x in corrective if x.get("actfinish")]:
            try:
                s = datetime.fromisoformat(w["reportdate"].replace("Z", "+00:00"))
                e = datetime.fromisoformat(w["actfinish"].replace("Z", "+00:00"))
                repair_times.append((e - s).total_seconds() / 3600)
            except Exception:
                pass
        mttr = round(sum(repair_times) / len(repair_times), 2) if repair_times else 0

        # MTBF
        period_hrs = period_months * 30 * 24
        total_downtime = sum(float(w.get("actlabhrs", 0) or 0) for w in corrective)
        mtbf = round((period_hrs - total_downtime) / len(corrective), 2) if corrective else period_hrs

        # Cost
        total_labor_cost = sum(float(w.get("actlabcost", 0) or 0) for w in wos)
        avg_wo_cost = round(total_labor_cost / len(wos), 2) if wos else 0

        # --- PM compliance ---
        # Single-condition WHERE only (siteid); status filter applied in Python.
        pm_result = await query_object_structure(
            entity="pm",
            filters={"site_id": site_id},
            select=["pmnum", "status", "nextduedate", "lastcompdate"],
            page_size=5,
        )
        pms_raw: List[Dict] = pm_result.get("data", []) if "error" not in pm_result else []
        pms: List[Dict] = [p for p in pms_raw if str(p.get("status", "")).upper() == "ACTIVE"]
        overdue_pms = [p for p in pms if p.get("nextduedate", "9999") < now_str]
        pm_compliance_pct = round(((len(pms) - len(overdue_pms)) / len(pms)) * 100, 1) if pms else 100

        # Priority breakdown
        priority_map: Dict[str, int] = {}
        for w in wos:
            p = str(w.get("priority", "N/A"))
            priority_map[p] = priority_map.get(p, 0) + 1

        duration_ms = int((time.monotonic() - start) * 1000)
        return _envelope(
            {
                "site_id": site_id,
                "period_months": period_months,
                "generated_at": datetime.now().isoformat(),
                "work_order_kpis": {
                    "total_wos": len(wos),
                    "completed": len(completed),
                    "backlog": len(backlog),
                    "overdue": len(overdue),
                    "corrective": len(corrective),
                    "priority_breakdown": priority_map,
                },
                "reliability_kpis": {
                    "mttr_hours": mttr,
                    "mtbf_hours": mtbf,
                    "total_downtime_hours": round(total_downtime, 2),
                    "availability_pct": round(((period_hrs - total_downtime) / period_hrs) * 100, 2),
                },
                "pm_kpis": {
                    "total_active_pms": len(pms),
                    "overdue_pms": len(overdue_pms),
                    "pm_compliance_pct": pm_compliance_pct,
                },
                "cost_kpis": {
                    "total_labor_cost": round(total_labor_cost, 2),
                    "avg_wo_cost": avg_wo_cost,
                },
            },
            duration_ms=duration_ms
        )
    except (MaximoAPIError, MaximoAuthError) as exc:
        return _error(str(exc))
    except Exception as exc:
        return _error(f"Unexpected error [{type(exc).__name__}]: {exc!r}", "INTERNAL_ERROR")


@require_role("supervisor")
async def export_workorders_excel(
    site_id: str,
    filters: Optional[Dict[str, str]] = None,
    max_records: int = 1000,
) -> Dict[str, Any]:
    """
    Generate an Excel workbook of work orders for a site and return it as base64.

    Args:
        site_id:     Site ID to export
        filters:     Optional OSLC where-clause fragments as {field: value}
        max_records: Maximum records to export (default: 1000)

    Returns:
        base64-encoded Excel file, filename, and record count.
    """
    if not site_id:
        return _error("site_id is required", "VALIDATION_ERROR")

    start = time.monotonic()

    try:
        import openpyxl  # type: ignore
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return _error("openpyxl is required: pip install openpyxl", "DEPENDENCY_ERROR")

    try:
        client = await get_connected_client()
        from core.oslc_utils import safe_field_name  # already imported above; guard for clarity

        # Single-condition WHERE on siteid; additional filters applied in Python
        # (compound WHERE with multiple equality clauses can drop the connection
        # on some Maximo builds — same constraint that drove Wave-2 refactors).
        params = client.build_oslc_query(
            where=f'siteid="{oslc_escape(site_id)}"',
            select="wonum,description,status,priority,assetnum,worktype,reportdate,actfinish,actlabhrs,actlabcost,siteid",
            order_by="-reportdate",
            page_size=max_records,
        )
        data = await client.get("/os/mxwo", params=params)
        all_wos: List[Dict] = data.get("member", [])

        # Validate filter field names + apply equality filters in Python
        validated_filters: Dict[str, str] = {}
        if filters:
            for field, value in filters.items():
                try:
                    safe_field_name(field)
                except ValueError:
                    continue
                validated_filters[field] = str(value)

        site_u = site_id.upper()
        wos: List[Dict] = []
        for wo in all_wos:
            if (wo.get("siteid") or "").upper() != site_u:
                continue
            keep = True
            for field, value in validated_filters.items():
                if str(wo.get(field) or "").upper() != value.upper():
                    keep = False
                    break
            if keep:
                wos.append(wo)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Work Orders"

        headers = ["WO Number", "Description", "Status", "Priority", "Asset", "Type",
                   "Report Date", "Actual Finish", "Actual Hours", "Labor Cost"]
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="1F4E79")

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill

        for row, wo in enumerate(wos, 2):
            ws.cell(row=row, column=1, value=wo.get("wonum", ""))
            ws.cell(row=row, column=2, value=wo.get("description", ""))
            ws.cell(row=row, column=3, value=wo.get("status", ""))
            ws.cell(row=row, column=4, value=wo.get("priority", ""))
            ws.cell(row=row, column=5, value=wo.get("assetnum", ""))
            ws.cell(row=row, column=6, value=wo.get("worktype", ""))
            ws.cell(row=row, column=7, value=wo.get("reportdate", ""))
            ws.cell(row=row, column=8, value=wo.get("actfinish", ""))
            ws.cell(row=row, column=9, value=wo.get("actlabhrs", ""))
            ws.cell(row=row, column=10, value=wo.get("actlabcost", ""))

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 50)

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()
        filename = f"workorders_{site_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

        duration_ms = int((time.monotonic() - start) * 1000)
        return _envelope(
            {"filename": filename, "base64_content": b64, "record_count": len(wos), "site_id": site_id},
            duration_ms=duration_ms
        )
    except (MaximoAPIError, MaximoAuthError) as exc:
        return _error(str(exc))
    except Exception as exc:
        return _error(f"Unexpected error [{type(exc).__name__}]: {exc!r}", "INTERNAL_ERROR")


@require_role("supervisor")
async def export_asset_report_pdf(
    site_id: str,
    asset_group: Optional[str] = None,
    max_records: int = 200,
) -> Dict[str, Any]:
    """
    Generate a PDF asset report for a site and return it as base64.

    Args:
        site_id:     Site ID to report on
        asset_group: Optional asset type/group filter
        max_records: Max assets to include

    Returns:
        base64-encoded PDF file, filename, and record count.
    """
    if not site_id:
        return _error("site_id is required", "VALIDATION_ERROR")

    start = time.monotonic()

    try:
        from reportlab.lib.pagesizes import A4  # type: ignore
        from reportlab.lib.styles import getSampleStyleSheet  # type: ignore
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer  # type: ignore
        from reportlab.lib import colors  # type: ignore
    except ImportError:
        return _error("reportlab is required: pip install reportlab", "DEPENDENCY_ERROR")

    try:
        client = await get_connected_client()
        # Single-condition WHERE on siteid; asset_group filter applied in Python.
        params = client.build_oslc_query(
            where=f'siteid="{oslc_escape(site_id)}"',
            select="assetnum,description,status,assettype,location,serialnum,installdate,siteid",
            order_by="+assetnum",
            page_size=max_records,
        )
        data = await client.get("/os/mxasset", params=params)
        all_assets: List[Dict] = data.get("member", [])
        site_u = site_id.upper()
        assets = [a for a in all_assets if (a.get("siteid") or "").upper() == site_u]
        if asset_group:
            ag_u = asset_group.upper()
            assets = [a for a in assets if (a.get("assettype") or "").upper() == ag_u]

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []

        story.append(Paragraph(f"Asset Report — Site: {site_id}", styles["Title"]))
        story.append(Paragraph(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}", styles["Normal"]))
        story.append(Spacer(1, 20))

        headers = ["Asset #", "Description", "Status", "Type", "Location", "Serial #"]
        table_data = [headers]
        for a in assets:
            table_data.append([
                a.get("assetnum", ""),
                (a.get("description", "") or "")[:40],
                a.get("status", ""),
                a.get("assettype", ""),
                a.get("location", ""),
                a.get("serialnum", ""),
            ])

        tbl = Table(table_data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1F4E79")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#EBF3FB")]),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
        ]))
        story.append(tbl)
        doc.build(story)

        buf.seek(0)
        b64 = base64.b64encode(buf.read()).decode()
        filename = f"assets_{site_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"

        duration_ms = int((time.monotonic() - start) * 1000)
        return _envelope(
            {"filename": filename, "base64_content": b64, "record_count": len(assets), "site_id": site_id},
            duration_ms=duration_ms
        )
    except (MaximoAPIError, MaximoAuthError) as exc:
        return _error(str(exc))
    except Exception as exc:
        return _error(f"Unexpected error [{type(exc).__name__}]: {exc!r}", "INTERNAL_ERROR")


async def generate_carbon_table(
    object_structure: str,
    data: List[Dict[str, Any]],
    columns: List[Dict[str, str]],
) -> Dict[str, Any]:
    """
    Render data as an IBM Carbon Design System HTML table.
    Compatible with existing Maximo MCP Carbon table conventions.

    Args:
        object_structure: Name of the object for table title
        data:             List of row dicts
        columns:          List of {key, header} dicts defining columns

    Returns:
        HTML string of the rendered Carbon table.
    """
    if not data or not columns:
        return _error("data and columns are required", "VALIDATION_ERROR")

    start = time.monotonic()
    headers_html = "".join(
        f'<th scope="col" class="bx--table-header-label">{col["header"]}</th>'
        for col in columns
    )
    rows_html = ""
    for row in data:
        cells = "".join(
            f'<td class="bx--table-column-menu">{row.get(col["key"], "")}</td>'
            for col in columns
        )
        rows_html += f"<tr>{cells}</tr>"

    html = f"""
<div class="bx--data-table-container">
  <div class="bx--data-table-header">
    <h4 class="bx--data-table-header__title">{object_structure}</h4>
  </div>
  <table class="bx--data-table bx--data-table--sort">
    <thead><tr>{headers_html}</tr></thead>
    <tbody>{rows_html}</tbody>
  </table>
  <p class="bx--table-toolbar-content">{len(data)} records</p>
</div>"""

    duration_ms = int((time.monotonic() - start) * 1000)
    return _envelope({"html": html, "record_count": len(data)}, duration_ms=duration_ms)


@require_role("readonly")
async def get_failure_pareto(
    site_id: str,
    asset_num: Optional[str] = None,
    period_months: int = 12,
    top_n: int = 10,
) -> Dict[str, Any]:
    """
    Pareto chart of failure codes — the top N failure codes by frequency,
    with running cumulative percentage. Answers "what 20% of failure modes
    cause 80% of work?".

    Implementation: single-condition OSLC fetch on assetnum (when given) or
    siteid; date and worktype filters applied in Python because compound
    WHERE clauses can drop the connection on some Maximo builds.

    Args:
        site_id:       Site to analyse
        asset_num:     Optional asset filter (narrows to one machine)
        period_months: Look-back window in months (default 12)
        top_n:         Number of failure codes to return (default 10)
    """
    if not site_id:
        return _error("site_id is required", "VALIDATION_ERROR")

    start = time.monotonic()
    cutoff = (datetime.now() - timedelta(days=period_months * 30)).strftime(
        "%Y-%m-%dT00:00:00+00:00"
    )

    try:
        client = await get_connected_client()
        # Use the most-selective single-condition filter we have.
        if asset_num:
            where = f'assetnum="{oslc_escape(asset_num)}"'
        else:
            where = f'siteid="{oslc_escape(site_id)}"'
        params = client.build_oslc_query(
            where=where,
            select="wonum,siteid,assetnum,worktype,failurecode,reportdate",
            order_by="-reportdate",
            page_size=200,
        )
        data = await client.get("/os/mxwo", params=params)
        rows: List[Dict] = data.get("member", [])

        site_u = site_id.upper()
        corrective = {"CM", "EM"}
        counts: Dict[str, int] = {}
        total_with_code = 0
        for w in rows:
            if (w.get("siteid") or "").upper() != site_u:
                continue
            if (w.get("worktype") or "").upper() not in corrective:
                continue
            if (w.get("reportdate") or "") < cutoff:
                continue
            fc = (w.get("failurecode") or "").strip()
            if not fc:
                continue
            counts[fc] = counts.get(fc, 0) + 1
            total_with_code += 1

        ranked = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[:top_n]
        cumulative = 0
        pareto = []
        for code, count in ranked:
            cumulative += count
            pct = round((count / total_with_code) * 100, 1) if total_with_code else 0
            cum_pct = round((cumulative / total_with_code) * 100, 1) if total_with_code else 0
            pareto.append(
                {
                    "failure_code": code,
                    "count": count,
                    "pct": pct,
                    "cumulative_pct": cum_pct,
                }
            )

        duration_ms = int((time.monotonic() - start) * 1000)
        return _envelope(
            {
                "site_id": site_id,
                "asset_num": asset_num,
                "period_months": period_months,
                "total_corrective_wos": sum(
                    1
                    for w in rows
                    if (w.get("siteid") or "").upper() == site_u
                    and (w.get("worktype") or "").upper() in corrective
                    and (w.get("reportdate") or "") >= cutoff
                ),
                "total_with_failure_code": total_with_code,
                "pareto": pareto,
            },
            duration_ms=duration_ms,
        )
    except (MaximoAPIError, MaximoAuthError) as exc:
        return _error(str(exc))
    except Exception as exc:
        return _error(f"Unexpected error [{type(exc).__name__}]: {exc!r}", "INTERNAL_ERROR")


@require_role("readonly")
async def get_bad_actor_assets(
    site_id: str,
    period_months: int = 12,
    top_n: int = 10,
) -> Dict[str, Any]:
    """
    Top-N "bad actor" assets — those with the most corrective work orders
    over the look-back window. Each row also reports total labour hours
    (a downtime proxy) so a planner can sort by cost rather than count.

    Args:
        site_id:       Site to analyse
        period_months: Look-back window in months (default 12)
        top_n:         Number of assets to return (default 10)
    """
    if not site_id:
        return _error("site_id is required", "VALIDATION_ERROR")

    start = time.monotonic()
    cutoff = (datetime.now() - timedelta(days=period_months * 30)).strftime(
        "%Y-%m-%dT00:00:00+00:00"
    )

    try:
        client = await get_connected_client()
        params = client.build_oslc_query(
            where=f'siteid="{oslc_escape(site_id)}"',
            select="wonum,siteid,assetnum,worktype,actlabhrs,actlabcost,reportdate",
            order_by="-reportdate",
            page_size=200,
        )
        data = await client.get("/os/mxwo", params=params)
        rows: List[Dict] = data.get("member", [])

        site_u = site_id.upper()
        corrective = {"CM", "EM"}
        counts: Dict[str, int] = {}
        hours: Dict[str, float] = {}
        cost: Dict[str, float] = {}
        for w in rows:
            if (w.get("siteid") or "").upper() != site_u:
                continue
            if (w.get("worktype") or "").upper() not in corrective:
                continue
            if (w.get("reportdate") or "") < cutoff:
                continue
            asset = (w.get("assetnum") or "").strip()
            if not asset:
                continue
            counts[asset] = counts.get(asset, 0) + 1
            hours[asset] = hours.get(asset, 0.0) + float(w.get("actlabhrs") or 0)
            cost[asset] = cost.get(asset, 0.0) + float(w.get("actlabcost") or 0)

        ranked = sorted(counts.items(), key=lambda kv: kv[1], reverse=True)[:top_n]
        bad_actors = [
            {
                "asset_num": asset,
                "corrective_wo_count": cnt,
                "total_labor_hours": round(hours.get(asset, 0.0), 2),
                "total_labor_cost": round(cost.get(asset, 0.0), 2),
            }
            for asset, cnt in ranked
        ]

        duration_ms = int((time.monotonic() - start) * 1000)
        return _envelope(
            {
                "site_id": site_id,
                "period_months": period_months,
                "bad_actors": bad_actors,
                "total_assets_with_corrective_wo": len(counts),
            },
            duration_ms=duration_ms,
        )
    except (MaximoAPIError, MaximoAuthError) as exc:
        return _error(str(exc))
    except Exception as exc:
        return _error(f"Unexpected error [{type(exc).__name__}]: {exc!r}", "INTERNAL_ERROR")
